from datetime import datetime
from statistics import median
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_from_directory, jsonify, send_file
from flask_login import login_required
from app import db
from app.models import Tender, Project, UnitPrice, Category, WorkType, Firmy
from sqlalchemy import func, or_
from app.forms import TenderForm, UnitPriceForm
from app.storage_service import get_storage_service # <-- Główny import serwisu
import os
from werkzeug.utils import secure_filename
import fitz  # PyMuPDF
import pdfplumber
import xlrd
import openpyxl
import io
import traceback
from sqlalchemy.orm import joinedload # Dodany import
import pandas as pd
import numpy as np

tenders_bp = Blueprint('tenders', __name__, template_folder='templates', url_prefix='/tenders')


@tenders_bp.route('/')
@login_required
def list_tenders():
    query = Tender.query
    form = TenderForm()

    # Filtrowanie
    id_firmy = request.args.get('id_firmy', type=int)
    id_projektu = request.args.get('id_projektu', type=int)

    if id_firmy:
        query = query.filter(Tender.id_firmy == id_firmy)
    if id_projektu:
        query = query.filter(Tender.id_projektu == id_projektu)

    tenders = query.order_by(Tender.data_otrzymania.desc()).all()
    projects = Project.query.order_by(Project.nazwa_projektu).all()
    
    return render_template('tenders_list.html', tenders=tenders, projects=projects, form=form, title='Oferty')

@tenders_bp.route('/<int:tender_id>')
@login_required
def tender_details(tender_id):
    tender = Tender.query.get_or_404(tender_id)
    return render_template('tender_details.html', tender=tender, title=f"Szczegóły oferty: {tender.nazwa_oferty}")

@tenders_bp.route('/download/<int:tender_id>')
@login_required
def download_file(tender_id):
    tender = Tender.query.get_or_404(tender_id)
    if not tender.storage_path:
        flash('Do tej oferty nie ma przypisanego pliku.', 'warning')
        return redirect(url_for('tenders.tender_details', tender_id=tender.id))
        
    try:
        storage_service = get_storage_service()
        file_stream = storage_service.download(tender.storage_path)
        return send_file(file_stream, download_name=tender.original_filename, mimetype=tender.file_type, as_attachment=True)
    except FileNotFoundError:
        flash(f'Plik nie został znaleziony w lokalizacji: {tender.storage_path}', 'danger')
        return redirect(url_for('tenders.tender_details', tender_id=tender.id))
    except Exception as e:
        current_app.logger.error(f"Błąd pobierania pliku (tender_id: {tender_id}): {e}")
        flash(f'Wystąpił błąd podczas pobierania pliku.', 'danger')
        return redirect(url_for('tenders.tender_details', tender_id=tender.id))


@tenders_bp.route('/display/<int:tender_id>')
@login_required
def display_file(tender_id):
    tender = Tender.query.get_or_404(tender_id)
    if not tender.storage_path:
        flash('Do tej oferty nie ma przypisanego pliku.', 'warning')
        return redirect(url_for('tenders.tender_details', tender_id=tender.id))

    try:
        storage_service = get_storage_service()
        file_stream = storage_service.download(tender.storage_path)
        return send_file(file_stream, mimetype=tender.file_type)
    except FileNotFoundError:
        flash(f'Plik nie został znaleziony w lokalizacji: {tender.storage_path}', 'danger')
        return redirect(url_for('tenders.tender_details', tender_id=tender.id))
    except Exception as e:
        current_app.logger.error(f"Błąd wyświetlania pliku (tender_id: {tender_id}): {e}")
        flash(f'Wystąpił błąd podczas wyświetlania pliku.', 'danger')
        return redirect(url_for('tenders.tender_details', tender_id=tender.id))


@tenders_bp.route('/<int:tender_id>/extract_data', methods=['GET', 'POST'])
@login_required
def extract_data(tender_id):
    tender = Tender.query.get_or_404(tender_id)
    extracted_text = ""
    table_data = []
    unit_price_form = UnitPriceForm()
    unit_price_form.id_oferty.data = tender.id

    if unit_price_form.validate_on_submit():
        try:
            work_type_id = unit_price_form.id_work_type.data
            work_type = WorkType.query.get(work_type_id)
            category_id = work_type.id_kategorii if work_type else None
            new_unit_price = UnitPrice(
                id_work_type=work_type_id,
                nazwa_roboty=work_type.name if work_type else None,
                jednostka_miary=unit_price_form.jednostka_miary.data,
                cena_jednostkowa=float(unit_price_form.cena_jednostkowa.data.replace(',', '.')),
                id_oferty=tender.id,
                id_kategorii=category_id,
                uwagi=unit_price_form.uwagi.data
            )
            db.session.add(new_unit_price)
            db.session.commit()
            flash('Pozycja cenowa została dodana pomyślnie!', 'success')
            return redirect(url_for('tenders.extract_data', tender_id=tender.id, from_submit=True))
        except Exception as e:
            db.session.rollback()
            flash(f'Wystąpił błąd podczas dodawania pozycji cenowej: {e}', 'danger')

    if request.method == 'GET':
        file_content = None
        if not tender.storage_path:
            flash("Brak pliku do przetworzenia.", "warning")
        else:
            # --- NOWA LOGIKA CACHE'OWANIA PLIKÓW ---
            local_file_path = os.path.join(current_app.instance_path, 'uploads', secure_filename(tender.original_filename))
            print(f"DEBUG: Sprawdzam lokalną ścieżkę: {local_file_path}")

            if os.path.exists(local_file_path):
                print(f"DEBUG: Plik znaleziony lokalnie. Wczytuję z: {local_file_path}")
                try:
                    with open(local_file_path, 'rb') as f:
                        file_content = io.BytesIO(f.read())
                except Exception as e:
                    flash(f"Nie udało się odczytać pliku z lokalnego cache: {e}", "danger")
            else:
                print(f"DEBUG: Pliku nie ma lokalnie. Pobieram z GCS: {tender.storage_path}")
                try:
                    storage_service = get_storage_service()
                    file_stream = storage_service.download(tender.storage_path)
                    file_bytes = file_stream.read()
                    
                    # Zapisz plik w lokalnym cache
                    with open(local_file_path, 'wb') as f:
                        f.write(file_bytes)
                    print(f"DEBUG: Plik zapisany w lokalnym cache: {local_file_path}")

                    # Użyj pobranej zawartości
                    file_content = io.BytesIO(file_bytes)

                except Exception as e:
                    flash(f"Nie udało się wczytać pliku z GCS: {e}", "danger")
            # --- KONIEC NOWEJ LOGIKI ---

        if file_content:
            file_content.seek(0)
            filename_lower = tender.original_filename.lower()
            is_image_file = False
            display_original_pdf = False

            try:
                if filename_lower.endswith('.pdf'):
                    pdf_extracted_successfully = False
                    try:
                        with pdfplumber.open(file_content) as pdf:
                            has_tables = False
                            for page in pdf.pages:
                                tables = page.extract_tables()
                                if tables:
                                    has_tables = True
                                    for table in tables:
                                        table_data.extend(table)
                            
                            if has_tables:
                                pdf_extracted_successfully = True
                            else:
                                extracted_text = "".join([page.extract_text() for page in pdf.pages if page.extract_text()])
                                if extracted_text.strip():
                                    pdf_extracted_successfully = True
                    except Exception as pdf_error:
                        file_content.seek(0)
                        with fitz.open(stream=file_content, filetype="pdf") as doc:
                            for page in doc:
                                extracted_text += page.get_text("text", sort=True)
                            if extracted_text.strip():
                                pdf_extracted_successfully = True
                    
                    if not pdf_extracted_successfully:
                        display_original_pdf = True

                elif filename_lower.endswith('.xlsx'):
                    workbook = openpyxl.load_workbook(file_content, data_only=True)
                    for sheet in workbook.worksheets:
                        merged_ranges = list(sheet.merged_cells.ranges)
                        for merged_range in merged_ranges:
                            sheet.unmerge_cells(str(merged_range))

                        for row in sheet.iter_rows():
                            row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                            if any(row_data):
                                table_data.append(row_data)

                elif filename_lower.endswith('.xls'):
                    workbook = xlrd.open_workbook(file_contents=file_content.read())
                    for sheet in workbook.sheets():
                        for row_idx in range(sheet.nrows):
                            row_data = []
                            for col_idx in range(sheet.ncols):
                                cell_value = sheet.cell_value(row_idx, col_idx)
                                for rlo, rhi, clo, chi in sheet.merged_cells:
                                    if rlo <= row_idx < rhi and clo <= col_idx < chi:
                                        cell_value = sheet.cell_value(rlo, clo)
                                        break
                                row_data.append(str(cell_value) if cell_value is not None else "")
                            if any(row_data):
                                table_data.append(row_data)
                
                elif filename_lower.endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp')):
                    is_image_file = True
                else:
                    flash(f'Nieobsługiwany typ pliku do ekstrakcji danych: "{tender.original_filename}" (typ MIME: {tender.file_type})', 'warning')

            except Exception as e:
                traceback.print_exc()
                flash(f'Wystąpił błąd podczas ekstrakcji danych: {e}', 'danger')
    
    return render_template('extract_helper.html', tender=tender, extracted_text=extracted_text, table_data=table_data, unit_price_form=unit_price_form, categories=Category.query.order_by(Category.nazwa_kategorii).all(), unit_prices=tender.unit_prices.all(), title="Ekstrakcja danych z oferty", is_image_file=is_image_file if 'is_image_file' in locals() else False, display_original_pdf=display_original_pdf if 'display_original_pdf' in locals() else False)

@tenders_bp.route('/<int:tender_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_tender(tender_id):
    tender = Tender.query.get_or_404(tender_id)
    form = TenderForm(obj=tender)
    
    if request.method == 'GET':
        form.delete_existing_file.data = False

    if form.validate_on_submit():
        storage_service = get_storage_service()
        
        # Obsługa usuwania istniejącego pliku
        if form.delete_existing_file.data and not form.plik_oferty.data:
            if tender.storage_path:
                try:
                    storage_service.delete(tender.storage_path)
                    tender.original_filename = None
                    tender.storage_path = None
                    tender.file_type = None
                    flash('Istniejący plik został usunięty.', 'info')
                except Exception as e:
                    flash(f"Błąd podczas usuwania pliku: {e}", "danger")

        # Obsługa nowego pliku
        if form.plik_oferty.data:
            plik = form.plik_oferty.data
            filename = secure_filename(plik.filename)
            
            # Usuń stary plik przed wgraniem nowego
            if tender.storage_path:
                try:
                    storage_service.delete(tender.storage_path)
                except Exception as e:
                    flash(f"Nie udało się usunąć starego pliku, ale kontynuowano wgrywanie nowego. Błąd: {e}", "warning")

            try:
                tender.storage_path = storage_service.upload(plik.stream, filename, plik.mimetype)
                tender.original_filename = filename
                tender.file_type = plik.mimetype
            except Exception as e:
                flash(f"Błąd podczas wgrywania nowego pliku: {e}", "danger")
                return redirect(url_for('tenders.edit_tender', tender_id=tender.id))

        tender.nazwa_oferty = form.nazwa_oferty.data
        tender.data_otrzymania = form.data_otrzymania.data
        tender.status = form.status.data
        tender.id_firmy = form.id_firmy.data
        tender.id_projektu = form.id_projektu.data if form.id_projektu.data else None
        
        db.session.commit()
        flash('Oferta została zaktualizowana.', 'success')
        return redirect(url_for('tenders.tender_details', tender_id=tender.id))
    
    return render_template('tender_form.html', form=form, tender=tender, title=f"Edycja oferty: {tender.nazwa_oferty}")


@tenders_bp.route('/<int:tender_id>/delete', methods=['POST'])
@login_required
def delete_tender(tender_id):
    tender = Tender.query.get_or_404(tender_id)
    
    if tender.storage_path:
        try:
            storage_service = get_storage_service()
            storage_service.delete(tender.storage_path)
        except Exception as e:
            flash(f"Nie udało się usunąć pliku powiązanego z ofertą, ale oferta zostanie usunięta. Błąd: {e}", "warning")

    db.session.delete(tender)
    db.session.commit()
    flash('Oferta została usunięta.', 'success')
    return redirect(url_for('tenders.list_tenders'))

@tenders_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new_tender():
    form = TenderForm()
    if form.validate_on_submit():
        if form.plik_oferty.data:
            plik = form.plik_oferty.data
            filename = secure_filename(plik.filename)

            try:
                storage_service = get_storage_service()
                storage_path = storage_service.upload(plik.stream, filename, plik.mimetype)

                nowa_oferta = Tender(
                    nazwa_oferty=form.nazwa_oferty.data,
                    data_otrzymania=form.data_otrzymania.data,
                    status=form.status.data,
                    id_firmy=form.id_firmy.data,
                    id_projektu=form.id_projektu.data if form.id_projektu.data else None,
                    original_filename=filename,
                    storage_path=storage_path, 
                    file_type=plik.mimetype
                )
                
                db.session.add(nowa_oferta)
                db.session.commit()

                flash('Nowa oferta została dodana pomyślnie!', 'success')
                return redirect(url_for('tenders.list_tenders'))

            except Exception as e:
                flash(f"Wystąpił błąd podczas przesyłania pliku: {e}", "danger")

        else:
            flash('Proszę załączyć plik oferty.', 'danger')
            
    return render_template('tender_form.html', form=form, title='Nowa Oferta')

@tenders_bp.route('/unit_prices')
@login_required
def list_all_unit_prices():
    PER_PAGE = 15 # Liczba elementów na stronę

    nazwa_roboty_filter = request.args.get('nazwa_roboty', type=int)
    kategoria_filter = request.args.get('kategoria', type=int)
    id_oferty_filter = request.args.get('id_oferty', type=int)
    id_firmy_filter = request.args.get('id_firmy', type=int)
    id_projektu_filter = request.args.get('id_projektu', type=int)
    page = request.args.get('page', 1, type=int) # Pobierz numer strony

    query = UnitPrice.query.join(WorkType).join(Tender).join(Firmy)

    if nazwa_roboty_filter:
        query = query.filter(UnitPrice.id_work_type == nazwa_roboty_filter)
    if kategoria_filter:
        query = query.filter(UnitPrice.id_kategorii == kategoria_filter)
    if id_oferty_filter:
        query = query.filter(UnitPrice.id_oferty == id_oferty_filter)
    if id_firmy_filter:
        query = query.filter(Tender.id_firmy == id_firmy_filter)
    if id_projektu_filter:
        query = query.filter(Tender.id_projektu == id_projektu_filter)

    # Paginacja
    pagination = query.order_by(UnitPrice.id.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    unit_prices = pagination.items # Elementy dla bieżącej strony

    work_types = WorkType.query.order_by(WorkType.name).all()
    categories = Category.query.order_by(Category.nazwa_kategorii).all()
    
    # Zoptymalizowane zapytanie do pobierania ofert do filtra
    all_tenders_for_filter = Tender.query.options(
        joinedload(Tender.firma),
        joinedload(Tender.project)
    ).order_by(Tender.data_otrzymania.desc()).all()

    # Formatowanie ofert dla listy rozwijanej filtra
    formatted_tenders_for_filter = []
    for t in all_tenders_for_filter:
        company_name = t.firma.nazwa_firmy if t.firma else "Brak firmy"
        if len(company_name) > 10:
            company_name = company_name[:10] + "..."

        project_info = "Brak projektu"
        if t.project:
            project_info = t.project.skrot if t.project.skrot else t.project.nazwa_projektu

        label = f"{t.nazwa_oferty} ({t.status}) - {company_name} ({project_info})"
        formatted_tenders_for_filter.append((t.id, label))

    firmy = Firmy.query.order_by(Firmy.nazwa_firmy).all()
    projects = Project.query.order_by(Project.nazwa_projektu).all()

    return render_template(
        'unit_prices_list.html',
        unit_prices=unit_prices,
        work_types=work_types,
        categories=categories,
        tenders=formatted_tenders_for_filter, # Przekazujemy sformatowane oferty
        firmy=firmy,
        projects=projects,
        selected_nazwa_roboty=nazwa_roboty_filter,
        selected_kategoria=kategoria_filter,
        selected_id_oferty=id_oferty_filter,
        selected_id_firmy=id_firmy_filter,
        selected_id_projektu=id_projektu_filter,
        pagination=pagination, # Przekazujemy obiekt paginacji
        title='Wszystkie pozycje cenowe'
    )

@tenders_bp.route('/unit_prices/analysis')
@login_required
def unit_prices_analysis():
    DEFAULT_TENDER_LIMIT = 10
    MAX_TENDER_LIMIT = 20

    # Pobieranie parametrów filtrowania
    category_filter = request.args.get('category', type=int)
    tender_ids_filter = request.args.getlist('tenders', type=int)
    status_filter = request.args.get('status', '')
    date_from_filter = request.args.get('date_from', '')
    date_to_filter = request.args.get('date_to', '')

    # Pobierz zakres dat dla walidacji i ustawień datapickera
    min_date, max_date = db.session.query(func.min(Tender.data_otrzymania), func.max(Tender.data_otrzymania)).one()

    # Walidacja dat po stronie serwera
    if date_from_filter:
        try:
            from_date = datetime.strptime(date_from_filter, '%Y-%m-%d').date()
            if min_date and from_date < min_date:
                flash(f"Data 'od' ({date_from_filter}) jest wcześniejsza niż najstarsza oferta ({min_date}). Filtr daty 'od' został zignorowany.", "warning")
                date_from_filter = ''
        except ValueError:
            date_from_filter = '' # Ignoruj niepoprawny format
    
    if date_to_filter:
        try:
            to_date = datetime.strptime(date_to_filter, '%Y-%m-%d').date()
            if max_date and to_date > max_date:
                flash(f"Data 'do' ({date_to_filter}) jest późniejsza niż najnowsza oferta ({max_date}). Filtr daty 'do' został zignorowany.", "warning")
                date_to_filter = ''
        except ValueError:
            date_to_filter = '' # Ignoruj niepoprawny format

    work_types_query = WorkType.query.order_by(WorkType.name)
    if category_filter:
        work_types_query = work_types_query.filter(WorkType.id_kategorii == category_filter)
    all_work_types = work_types_query.all()

    tenders_truncated = False
    
    base_tenders_query = Tender.query.options(
        joinedload(Tender.firma),
        joinedload(Tender.project)
    ).order_by(Tender.data_otrzymania.desc())

    if status_filter:
        base_tenders_query = base_tenders_query.filter(Tender.status == status_filter)
    if date_from_filter:
        base_tenders_query = base_tenders_query.filter(Tender.data_otrzymania >= date_from_filter)
    if date_to_filter:
        base_tenders_query = base_tenders_query.filter(Tender.data_otrzymania <= date_to_filter)

    if tender_ids_filter:
        all_tenders = base_tenders_query.filter(Tender.id.in_(tender_ids_filter)).all()
        if len(all_tenders) > MAX_TENDER_LIMIT:
            all_tenders = all_tenders[:MAX_TENDER_LIMIT]
            tenders_truncated = True
            flash(f'Wybrano więcej niż {MAX_TENDER_LIMIT} ofert. Wyświetlono tylko {MAX_TENDER_LIMIT} najnowszych z wybranych.', 'info')
    else:
        all_tenders = base_tenders_query.limit(DEFAULT_TENDER_LIMIT).all()
        if not any([status_filter, date_from_filter, date_to_filter, tender_ids_filter]):
             flash(f'Domyślnie wyświetlono {DEFAULT_TENDER_LIMIT} najnowszych ofert. Użyj filtrów, aby wybrać inne.', 'info')

    formatted_tender_headers = {}
    SHORTEN_COMPANY_NAME_THRESHOLD = 5 

    for tender in all_tenders:
        company_name = tender.firma.nazwa_firmy if tender.firma else "Brak firmy"
        
        if len(all_tenders) > SHORTEN_COMPANY_NAME_THRESHOLD and len(company_name) > 10:
            company_name = company_name[:10] + "..."

        project_info = "Brak projektu"
        if tender.project:
            project_info = tender.project.skrot if tender.project.skrot else tender.project.nazwa_projektu

        header_label = f"{company_name} ({project_info})"
        formatted_tender_headers[tender.id] = header_label

    unit_prices_query = UnitPrice.query.join(WorkType).join(Tender)
    if category_filter:
        unit_prices_query = unit_prices_query.filter(UnitPrice.id_kategorii == category_filter)
    
    if all_tenders:
        unit_prices_query = unit_prices_query.filter(UnitPrice.id_oferty.in_([t.id for t in all_tenders]))
    else:
        unit_prices_query = unit_prices_query.filter(False)

    all_unit_prices = unit_prices_query.all()

    prices_table = {}
    for up in all_unit_prices:
        if up.id_work_type not in prices_table:
            prices_table[up.id_work_type] = {}
        if up.id_oferty not in prices_table[up.id_work_type]:
            prices_table[up.id_work_type][up.id_oferty] = []

        prices_table[up.id_work_type][up.id_oferty].append({
            'cena': up.cena_jednostkowa,
            'uwagi': up.uwagi or ""
        })

    categories = Category.query.order_by(Category.nazwa_kategorii).all()
    all_available_tenders = Tender.query.order_by(Tender.nazwa_oferty).all()
    
    all_statuses = [s[0] for s in db.session.query(Tender.status).distinct().order_by(Tender.status).all()]

    return render_template(
        'unit_prices_analysis.html',
        all_work_types=all_work_types,
        all_tenders=all_tenders,
        prices_table=prices_table,
        categories=categories,
        all_available_tenders=all_available_tenders,
        selected_category=category_filter,
        selected_tenders=tender_ids_filter,
        formatted_tender_headers=formatted_tender_headers,
        tenders_truncated=tenders_truncated,
        all_statuses=all_statuses,
        selected_status=status_filter,
        selected_date_from=date_from_filter,
        selected_date_to=date_to_filter,
        min_date=min_date.strftime('%Y-%m-%d') if min_date else '',
        max_date=max_date.strftime('%Y-%m-%d') if max_date else '',
        title='Porównanie cen jednostkowych'
    )

@tenders_bp.route('/unit_prices/analysis/time_series/<int:work_type_id>')
@login_required
def unit_prices_time_series_data(work_type_id):
    time_series_data = db.session.query(
        func.to_char(Tender.data_otrzymania, 'YYYY-MM'),
        func.avg(UnitPrice.cena_jednostkowa)
    ).join(Tender).filter(UnitPrice.id_work_type == work_type_id).group_by(func.to_char(Tender.data_otrzymania, 'YYYY-MM')).order_by(func.to_char(Tender.data_otrzymania, 'YYYY-MM')).all()

    labels = [row[0] for row in time_series_data]
    data = [float(row[1]) for row in time_series_data]

    return jsonify({'labels': labels, 'data': data})

@tenders_bp.route('/unit_prices/analysis/by_contractor/<int:work_type_id>')
@login_required
def unit_prices_by_contractor_data(work_type_id):
    contractor_data = db.session.query(
        Firmy.nazwa_firmy,
        func.avg(UnitPrice.cena_jednostkowa)
    ).select_from(UnitPrice).join(Tender).join(Firmy).filter(UnitPrice.id_work_type == work_type_id).group_by(Firmy.nazwa_firmy).order_by(Firmy.nazwa_firmy).all()

    labels = [row[0] for row in contractor_data]
    data = [float(row[1]) for row in contractor_data]

    return jsonify({'labels': labels, 'data': data})

@tenders_bp.route('/unit_prices/new_global', methods=['GET', 'POST'])
@login_required
def new_global_unit_price():
    form = UnitPriceForm()
    if form.validate_on_submit():
        try:
            work_type_id = form.id_work_type.data
            work_type = WorkType.query.get(work_type_id)
            category_id = work_type.id_kategorii if work_type else None
            new_unit_price = UnitPrice(
                id_work_type=work_type_id,
                nazwa_roboty=work_type.name if work_type else None,
                jednostka_miary=form.jednostka_miary.data,
                cena_jednostkowa=form.cena_jednostkowa.data,
                id_oferty=form.id_oferty.data,
                id_kategorii=category_id,
                uwagi=form.uwagi.data
            )
            db.session.add(new_unit_price)
            db.session.commit()
            flash('Pozycja cenowa została dodana pomyślnie!', 'success')
            return redirect(url_for('tenders.list_all_unit_prices'))
        except Exception as e:
            db.session.rollback()
            flash(f'Wystąpił błąd podczas dodawania pozycji cenowej: {e}', 'danger')

    return render_template('unit_price_form.html', form=form, title='Dodaj nową pozycję cenową', show_tender_select=True, category_field_always_disabled_unless_auto_filled=True)

@tenders_bp.route('/unit_price/<int:price_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_unit_price(price_id):
    price = UnitPrice.query.get_or_404(price_id)
    form = UnitPriceForm(obj=price)
    if form.validate_on_submit():
        price.id_work_type = form.id_work_type.data
        work_type = WorkType.query.get(price.id_work_type)
        price.nazwa_roboty = work_type.name if work_type else None
        price.jednostka_miary = form.jednostka_miary.data
        price.cena_jednostkowa = form.cena_jednostkowa.data
        price.id_kategorii = work_type.id_kategorii if work_type else None
        price.uwagi = form.uwagi.data
        db.session.commit()
        flash('Pozycja cenowa została zaktualizowana.', 'success')
        return redirect(url_for('tenders.tender_details', tender_id=price.id_oferty))

    return render_template('unit_price_form.html', form=form, title='Edycja pozycji cenowej', tender_id=price.id_oferty, category_field_always_disabled_unless_auto_filled=True)

@tenders_bp.route('/unit_price/<int:price_id>/delete', methods=['POST'])
@login_required
def delete_unit_price(price_id):
    price = UnitPrice.query.get_or_404(price_id)
    tender_id = price.id_oferty
    db.session.delete(price)
    db.session.commit()
    flash('Pozycja cenowa została usunięta.', 'success')
    return redirect(url_for('tenders.tender_details', tender_id=tender_id))

@tenders_bp.route('/analysis_dashboard')
@login_required
def analysis_dashboard():
    """
    Wyświetla pulpit analityczny do interaktywnej analizy cen jednostkowych.
    """
    work_types = WorkType.query.order_by(WorkType.name).all()
    selected_work_type_id = request.args.get('work_type_id', type=int)
    included_ids = request.args.getlist('include', type=int)
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    
    # Dodaj parametr paginacji
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Liczba rekordów na stronę

    date_from = None
    date_to = None
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Nieprawidłowy format daty 'od'. Użyj formatu RRRR-MM-DD.", "warning")
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Nieprawidłowy format daty 'do'. Użyj formatu RRRR-MM-DD.", "warning")

    # Pobierz zakres dat dla walidacji i ustawień datapickera
    min_tender_date, max_tender_date = db.session.query(func.min(Tender.data_otrzymania), func.max(Tender.data_otrzymania)).one()

    # Walidacja dat po stronie serwera
    if date_from and min_tender_date and date_from < min_tender_date:
        flash(f"Data 'od' ({date_from_str}) jest wcześniejsza niż najstarsza oferta ({min_tender_date}). Filtr daty 'od' został zignorowany.", "warning")
        date_from = None
        date_from_str = ''
    
    if date_to and max_tender_date and date_to > max_tender_date:
        flash(f"Data 'do' ({date_to_str}) jest późniejsza niż najnowsza oferta ({max_tender_date}). Filtr daty 'do' został zignorowany.", "warning")
        date_to = None
        date_to_str = ''

    stats = {}
    source_data_pagination = None  # Zmieniamy na obiekt paginacji
    if selected_work_type_id:
        # Pobierz wszystkie pozycje dla wybranego work_type_id (do wyświetlenia w tabeli)
        base_query_source = (
            db.session.query(
                UnitPrice.id,
                UnitPrice.cena_jednostkowa,
                UnitPrice.jednostka_miary,
                Tender.data_otrzymania,
                Firmy.nazwa_firmy,
                Tender.nazwa_oferty,
                Tender.id.label('tender_id'),
                Project.skrot.label('project_skrot'),
                UnitPrice.uwagi
            )
            .join(Tender, UnitPrice.id_oferty == Tender.id)
            .join(Firmy, Tender.id_firmy == Firmy.id_firmy)
            .outerjoin(Project, Tender.id_projektu == Project.id)
            .filter(UnitPrice.id_work_type == selected_work_type_id)
        )

        if date_from:
            base_query_source = base_query_source.filter(Tender.data_otrzymania >= date_from)
        if date_to:
            base_query_source = base_query_source.filter(Tender.data_otrzymania <= date_to)

        # Zastępujemy .all() paginacją
        source_data_pagination = base_query_source.order_by(Tender.data_otrzymania.desc()).paginate(page=page, per_page=per_page, error_out=False)
        
        # Jeśli nie ma included_ids, domyślnie zaznacz pozycje bez uwag
        if not included_ids and source_data_pagination.items:
            included_ids = [row.id for row in source_data_pagination.items if not row.uwagi]
        
        # Oblicz statystyki na podstawie wybranych pozycji
        if included_ids:
            base_query_stats = (db.session.query(UnitPrice.cena_jednostkowa)
                                .join(Tender, UnitPrice.id_oferty == Tender.id)
                                .filter(UnitPrice.id_work_type == selected_work_type_id)
                                .filter(UnitPrice.id.in_(included_ids)))

            if date_from:
                base_query_stats = base_query_stats.filter(Tender.data_otrzymania >= date_from)
            if date_to:
                base_query_stats = base_query_stats.filter(Tender.data_otrzymania <= date_to)

            prices = base_query_stats.all()
            price_values = [float(p[0]) for p in prices]
            
            if price_values:
                stats = {
                    'min_price': min(price_values),
                    'max_price': max(price_values),
                    'avg_price': sum(price_values) / len(price_values),
                    'median_price': median(price_values),
                    'offer_count': len(price_values)
                }

    return render_template('analysis_dashboard.html', 
                           title='Pulpit Analityczny Cen',
                           work_types=work_types,
                           selected_work_type_id=selected_work_type_id,
                           stats=stats,
                           source_data=source_data_pagination,  # Teraz przekazujemy obiekt paginacji
                           included_ids=included_ids,
                           date_from=date_from_str,
                           date_to=date_to_str,
                           min_tender_date=min_tender_date.strftime('%Y-%m-%d') if min_tender_date else '',
                           max_tender_date=max_tender_date.strftime('%Y-%m-%d') if max_tender_date else '')

def _get_filtered_data_as_df(work_type_id):
    """
    Funkcja pomocnicza do pobierania i filtrowania danych cenowych jako DataFrame Pandas.
    """
    included_ids = request.args.getlist('include', type=int)
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    date_from = None
    date_to = None
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Podstawowe zapytanie
    query = (
        db.session.query(
            UnitPrice.cena_jednostkowa,
            Tender.data_otrzymania,
            Firmy.nazwa_firmy,
            UnitPrice.id,
            UnitPrice.uwagi
        )
        .join(Tender, UnitPrice.id_oferty == Tender.id)
        .join(Firmy, Tender.id_firmy == Firmy.id_firmy)
        .filter(UnitPrice.id_work_type == work_type_id)
    )

    # Filtrowanie po dacie
    if date_from:
        query = query.filter(Tender.data_otrzymania >= date_from)
    if date_to:
        query = query.filter(Tender.data_otrzymania <= date_to)
    
    # Wczytanie danych do DataFrame
    df = pd.read_sql(query.statement, db.get_engine())
    if df.empty:
        return pd.DataFrame()

    # Konwersja typów
    df['cena_jednostkowa'] = pd.to_numeric(df['cena_jednostkowa'])
    df['data_otrzymania'] = pd.to_datetime(df['data_otrzymania'])

    # Filtrowanie po ID (jeśli podane)
    if not included_ids:
        # Domyślnie użyj pozycji bez uwag
        df_filtered = df[df['uwagi'].isnull() | (df['uwagi'] == '')].copy()
    else:
        df_filtered = df[df['id'].isin(included_ids)].copy()

    return df_filtered

@tenders_bp.route('/api/price_evolution/<int:work_type_id>')
@login_required
def price_evolution_data(work_type_id):
    """
    Zwraca dane do wykresu ewolucji ceny w czasie dla danego rodzaju roboty.
    """
    df = _get_filtered_data_as_df(work_type_id)
    if df.empty:
        return jsonify({'labels': [], 'values': []})

    df.set_index('data_otrzymania', inplace=True)
    monthly_avg = df.resample('ME')['cena_jednostkowa'].mean().dropna()
    
    labels = monthly_avg.index.strftime('%Y-%m').tolist()
    values = monthly_avg.values.tolist()
    
    return jsonify({'labels': labels, 'values': values})

@tenders_bp.route('/api/price_by_contractor/<int:work_type_id>')
@login_required
def price_by_contractor_data(work_type_id):
    """
    Zwraca dane do wykresu porównania cen wg wykonawcy dla danego rodzaju roboty.
    """
    df = _get_filtered_data_as_df(work_type_id)
    if df.empty:
        return jsonify({'labels': [], 'values': []})

    avg_by_contractor = df.groupby('nazwa_firmy')['cena_jednostkowa'].mean().sort_values(ascending=False)
    
    labels = avg_by_contractor.index.tolist()
    values = avg_by_contractor.values.tolist()

    return jsonify({'labels': labels, 'values': values})

@tenders_bp.route('/api/price_distribution/<int:work_type_id>')
@login_required
def price_distribution_data(work_type_id):
    """
    Zwraca dane do wykresu rozkładu cen (histogramu) dla danego rodzaju roboty.
    """
    df = _get_filtered_data_as_df(work_type_id)
    if df.empty:
        return jsonify({'labels': [], 'values': []})

    prices = df['cena_jednostkowa'].dropna().tolist()
    if not prices:
        return jsonify({'labels': [], 'values': []})

    min_price = min(prices)
    max_price = max(prices)
    num_bins = 7

    if min_price == max_price:
        return jsonify({'labels': [f'{min_price:.2f} zł'], 'values': [len(prices)]})

    # Użyj pd.cut do stworzenia przedziałów
    bins = pd.cut(prices, bins=num_bins)
    bin_counts = bins.value_counts().sort_index()

    labels = [str(interval) for interval in bin_counts.index]
    values = bin_counts.values.tolist()

    return jsonify({'labels': labels, 'values': values})

@tenders_bp.route('/api/price_trends/<int:work_type_id>')
@login_required
def price_trends_data(work_type_id):
    """
    Zwraca dane do wykresu trendu cenowego wraz z linią trendu.
    """
    df = _get_filtered_data_as_df(work_type_id)
    if df.empty or len(df) < 2:
        return jsonify({'labels': [], 'datasets': []})

    df = df.sort_values('data_otrzymania')
    
    # Konwertuj daty na wartości numeryczne (liczba dni od daty minimalnej)
    x_numeric = (df['data_otrzymania'] - df['data_otrzymania'].min()).dt.days
    y = df['cena_jednostkowa']

    # Oblicz linię trendu (regresja liniowa)
    coeffs = np.polyfit(x_numeric, y, 1)
    trend_line = np.poly1d(coeffs)
    
    # Przygotuj dane do wykresu
    scatter_data = [{'x': row['data_otrzymania'].strftime('%Y-%m-%d'), 'y': row['cena_jednostkowa']} for index, row in df.iterrows()]
    trend_data = [{'x': df['data_otrzymania'].min().strftime('%Y-%m-%d'), 'y': trend_line(x_numeric.min())},
                  {'x': df['data_otrzymania'].max().strftime('%Y-%m-%d'), 'y': trend_line(x_numeric.max())}]

    return jsonify({
        'scatter_data': scatter_data,
        'trend_data': trend_data
    })

@tenders_bp.route('/api/price_seasonality/<int:work_type_id>')
@login_required
def price_seasonality_data(work_type_id):
    """
    Zwraca dane do wykresu sezonowości cen (średnia cena w każdym miesiącu).
    """
    df = _get_filtered_data_as_df(work_type_id)
    if df.empty:
        return jsonify({'labels': [], 'values': []})

    df['month'] = df['data_otrzymania'].dt.month
    seasonality = df.groupby('month')['cena_jednostkowa'].mean()
    
    # Upewnij się, że mamy wszystkie 12 miesięcy
    seasonality = seasonality.reindex(range(1, 13))

    month_names = ['Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec', 'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']
    
    labels = [month_names[i-1] for i in seasonality.index]
    # Zamień NaN na None (który jest konwertowany na null w JSON)
    values = [value if not np.isnan(value) else None for value in seasonality.values]

    return jsonify({'labels': labels, 'values': values})

# @tenders_bp.route('/api/contractor_competitiveness/<int:work_type_id>')
# @login_required
# def contractor_competitiveness_data(work_type_id):
#     """
#     Zwraca dane do wykresu pudełkowego (box plot) konkurencyjności wykonawców.
#     """
#     df = _get_filtered_data_as_df(work_type_id)
#     if df.empty:
#         return jsonify({'labels': [], 'datasets': []})

#     # Grupuj po wykonawcy i oblicz statystyki dla box plot
#     grouped = df.groupby('nazwa_firmy')['cena_jednostkowa'].apply(list)
    
#     # Filtruj wykonawców z małą liczbą ofert, aby wykres był czytelny
#     grouped = grouped[grouped.apply(len) >= 2]
    
#     if grouped.empty:
#         return jsonify({'labels': [], 'datasets': []})

#     # Sortuj wg mediany ceny
#     sorted_labels = grouped.apply(np.median).sort_values().index
    
#     # Przygotuj dane w formacie dla chartjs-chart-box-and-violin-plot
#     # WAŻNE: Dane dla każdego wykonawcy muszą być posortowane rosnąco
#     boxplot_data = [sorted(grouped[label]) for label in sorted_labels]

#     return jsonify({
#         'labels': sorted_labels.tolist(),
#         'datasets': [{
#             'label': 'Rozkład cen',
#             'data': boxplot_data
#         }]
#     })

# @tenders_bp.route('/api/contractor_stats/<int:work_type_id>')
# @login_required
# def contractor_stats_data(work_type_id):
#     """
#     Zwraca dane do prostego wykresu konkurencyjności wykonawców 
#     z dodatkowymi statystykami w tooltipie.
#     """
#     df = _get_filtered_data_as_df(work_type_id)
#     if df.empty:
#         return jsonify({'labels': [], 'avg_values': [], 'stats': []})

#     # Grupuj po wykonawcy i oblicz statystyki
#     grouped = df.groupby('nazwa_firmy')['cena_jednostkowa'].agg(['mean', 'min', 'max', 'count'])
    
#     # Filtruj wykonawców z małą liczbą ofert
#     grouped = grouped[grouped['count'] >= 2]
    
#     if grouped.empty:
#         return jsonify({'labels': [], 'avg_values': [], 'stats': []})

#     # Sortuj wg średniej ceny
#     grouped = grouped.sort_values('mean')
    
#     # Przygotuj dane
#     stats = []
#     for _, row in grouped.iterrows():
#         stats.append({
#             'avg': float(row['mean']),
#             'min': float(row['min']),
#             'max': float(row['max']),
#             'count': int(row['count'])
#         })
    
#     return jsonify({
#         'labels': grouped.index.tolist(),
#         'avg_values': grouped['mean'].round(2).tolist(),
#         'stats': stats
#     })

@tenders_bp.route('/api/contractor_competitiveness_alternative/<int:work_type_id>')
@login_required
def contractor_competitiveness_alternative_data(work_type_id):
    """
    Zwraca dane do alternatywnego wykresu konkurencyjności wykonawców 
    (słupkowy z min, max, średnia zamiast box plot).
    """
    df = _get_filtered_data_as_df(work_type_id)
    if df.empty:
        return jsonify({'labels': [], 'avg_values': [], 'min_values': [], 'max_values': []})

    # Grupuj po wykonawcy i oblicz statystyki
    grouped = df.groupby('nazwa_firmy')['cena_jednostkowa'].agg(['mean', 'min', 'max', 'count'])
    
    # Filtruj wykonawców z małą liczbą ofert
    grouped = grouped[grouped['count'] >= 2]
    
    if grouped.empty:
        return jsonify({'labels': [], 'avg_values': [], 'min_values': [], 'max_values': []})

    # Sortuj wg średniej ceny
    grouped = grouped.sort_values('mean')
    
    return jsonify({
        'labels': grouped.index.tolist(),
        'avg_values': grouped['mean'].round(2).tolist(),
        'min_values': grouped['min'].round(2).tolist(),
        'max_values': grouped['max'].round(2).tolist()
    })

# WAŻNE: To jest endpoint tylko do debugowania! Pamiętaj, aby go usunąć po zakończeniu testów.
@tenders_bp.route('/debug/list_files')
def list_uploaded_files():
    # Importujemy 'os' i 'current_app' jeśli nie ma ich w zasięgu
    import os
    from flask import current_app

    # Używamy current_app.instance_path, co jest bardziej niezawodne niż odwołanie do config.
    # To zawsze wskazuje na folder 'instance'.
    upload_folder = os.path.join(current_app.instance_path, 'uploads')
    
    # Sprawdzamy, czy katalog w ogóle istnieje
    if not os.path.exists(upload_folder):
        return f"Katalog docelowy nie istnieje: {upload_folder}"

    # Listujemy pliki w tym katalogu
    try:
        files = os.listdir(upload_folder)
        if not files:
            return f"Katalog jest pusty: {upload_folder}"
        else:
            # Zwracamy listę plików jako prosty tekst
            return f"Pliki w katalogu {upload_folder}:<br>" + "<br>".join(files)
    except Exception as e:
        return f"Wystąpił błąd podczas listowania plików: {e}"
